from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Response, Request
from pydantic import BaseModel
import os
import requests
import json
from fastapi.middleware.cors import CORSMiddleware
import sys
import shutil
import tempfile
from typing import Optional, AsyncGenerator
import io
from fastapi.responses import StreamingResponse
import aiohttp
import asyncio
import uuid
import pathlib

# AQUI DEFINIMOS EXPLICITAMENTE AS VARIÁVEIS ANTES DE QUALQUER CARREGAMENTO
# Configuração FIXA do Ollama - SEMPRE vai usar esses valores independente do .env
OLLAMA_API_BASE = "http://localhost:11434"
OLLAMA_API_URL = f"{OLLAMA_API_BASE}/api/generate"
OLLAMA_MODEL = "deepseek-r1:32b"  # Forçado explicitamente - NUNCA será sobrescrito

# DEPOIS carregamos .env apenas para outras variáveis
from dotenv import load_dotenv

# Bibliotecas para processar diferentes tipos de arquivos
try:
    import PyPDF2  # Para arquivos PDF
    import docx    # Para arquivos DOCX
    import openpyxl # Para arquivos XLSX
    LIBS_DISPONÍVEIS = True
except ImportError:
    LIBS_DISPONÍVEIS = False
    print("Aviso: Algumas bibliotecas para processamento de arquivos não estão instaladas.")
    print("Para usar todas as funcionalidades, instale: pip install pypdf2 python-docx openpyxl")

# Tentar carregar variáveis de ambiente de múltiplos locais possíveis
ENV_PATHS = [
    ".env",                                           # No diretório atual
    "../.env",                                        # No diretório pai
    os.path.join(os.path.dirname(__file__), "../.env"), # No diretório pai relativo a este arquivo
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "../.env"), # Caminho absoluto
    "/mnt/projetos_nexor/Servidor_Nexortech/Projetos/GPTPOL_OLLAMA/.env"  # Caminho completo
]

env_loaded = False
for env_path in ENV_PATHS:
    if os.path.exists(env_path):
        print(f"Carregando variáveis de ambiente de: {env_path}")
        load_dotenv(env_path)
        env_loaded = True
        break

if not env_loaded:
    print("Aviso: Arquivo .env não encontrado. Usando valores padrão.")

print(f"URL da API Ollama: {OLLAMA_API_URL}")
print(f"Modelo Ollama: {OLLAMA_MODEL}")

# App FastAPI
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Permitir todas as origens
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuração da pasta para uploads temporários
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
print(f"Diretório de uploads: {UPLOAD_DIR}")

# Modelo da requisição
class Pergunta(BaseModel):
    pergunta: str
    thinkknMode: bool = False

# Rota raiz
@app.get("/")
async def root():
    return {"message": "API GPTPOL está ativa."}

# Função para extrair texto de um arquivo PDF
def extrair_texto_pdf(arquivo):
    try:
        pdf_reader = PyPDF2.PdfReader(arquivo)
        texto = ""
        for pagina in range(len(pdf_reader.pages)):
            texto += pdf_reader.pages[pagina].extract_text() + "\n"
        return texto
    except Exception as e:
        print(f"Erro ao processar PDF: {e}")
        return f"Erro ao processar PDF: {str(e)}"

# Função para extrair texto de um arquivo DOCX
def extrair_texto_docx(arquivo):
    try:
        doc = docx.Document(arquivo)
        texto = ""
        for paragrafo in doc.paragraphs:
            texto += paragrafo.text + "\n"
        return texto
    except Exception as e:
        print(f"Erro ao processar DOCX: {e}")
        return f"Erro ao processar DOCX: {str(e)}"

# Função para extrair texto de um arquivo XLSX
def extrair_texto_xlsx(arquivo):
    try:
        workbook = openpyxl.load_workbook(arquivo)
        texto = ""
        for planilha in workbook.sheetnames:
            sheet = workbook[planilha]
            texto += f"Planilha: {planilha}\n"
            for linha in sheet.iter_rows(values_only=True):
                texto += " | ".join([str(celula) if celula is not None else "" for celula in linha]) + "\n"
            texto += "\n"
        return texto
    except Exception as e:
        print(f"Erro ao processar XLSX: {e}")
        return f"Erro ao processar XLSX: {str(e)}"

# Função para processar arquivos enviados
def processar_arquivo(arquivo, tipo_arquivo):
    if tipo_arquivo == "application/pdf":
        return extrair_texto_pdf(arquivo)
    elif tipo_arquivo in ["application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
        return extrair_texto_docx(arquivo)
    elif tipo_arquivo in ["application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
        return extrair_texto_xlsx(arquivo)
    elif tipo_arquivo == "text/plain":
        return arquivo.read().decode("utf-8")
    else:
        return "Tipo de arquivo não suportado."

# Função para gerar respostas via Ollama
def gerar_resposta_ollama(prompt):
    try:
        payload = {
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False,
            "options": {
                "num_ctx": 4096,           # Contexto para DeepSeek 14B
                "num_thread": 8,           # Reduzido para adequar ao modelo menor
                "temperature": 0.7,        # Temperatura padrão
                "top_k": 40,               # Valor padrão
                "top_p": 0.9,              # Valor padrão
                "num_predict": 512,        # Tamanho razoável para resposta
                "repeat_penalty": 1.1,     # Penalidade leve para repetição
                "stop": ["<|end|>", "<|user|>"]  # Tokens de fim específicos para DeepSeek
            }
        }
        
        print(f"Enviando requisição ao modelo {OLLAMA_MODEL}")
        # Aumentando timeout para 60 segundos (1 minuto) - modelo menor é mais rápido
        response = requests.post(OLLAMA_API_URL, json=payload, timeout=60)
        
        if response.status_code == 200:
            result = response.json()
            return result.get("response", "Não foi possível gerar uma resposta.")
        else:
            print(f"Erro Ollama: Status {response.status_code}")
            raise HTTPException(status_code=response.status_code, detail=f"Erro ao conectar com o modelo: {response.status_code}")
    except requests.exceptions.Timeout:
        print("Timeout ao chamar o modelo Ollama.")
        raise HTTPException(status_code=408, detail="Tempo limite excedido. O modelo está demorando mais do que o esperado. Por favor, tente uma pergunta mais simples ou tente novamente mais tarde.")
    except Exception as e:
        print(f"Erro ao conectar com Ollama: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao conectar com o modelo: {str(e)}")

# Rota para upload de arquivos
@app.post("/api/upload")
async def upload_arquivo(file: UploadFile = File(...)):
    try:
        # Verificar se as bibliotecas necessárias estão disponíveis
        if not LIBS_DISPONÍVEIS:
            raise HTTPException(status_code=500, detail="Bibliotecas para processamento de arquivos não estão instaladas no servidor.")
        
        # Verificar tipo de arquivo
        tipo_arquivo = file.content_type
        tipos_permitidos = [
            "application/pdf", 
            "application/msword", 
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "text/plain",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ]
        
        if tipo_arquivo not in tipos_permitidos:
            raise HTTPException(status_code=400, detail=f"Tipo de arquivo não suportado: {tipo_arquivo}")
        
        # Criar arquivo temporário para processar
        with tempfile.NamedTemporaryFile(delete=False) as temp:
            shutil.copyfileobj(file.file, temp)
            temp_path = temp.name
        
        try:
            # Processar o arquivo baseado no tipo
            if tipo_arquivo == "application/pdf":
                with open(temp_path, "rb") as f:
                    texto = extrair_texto_pdf(f)
            elif tipo_arquivo in ["application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
                texto = extrair_texto_docx(temp_path)
            elif tipo_arquivo in ["application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                texto = extrair_texto_xlsx(temp_path)
            elif tipo_arquivo == "text/plain":
                with open(temp_path, "r", encoding="utf-8") as f:
                    texto = f.read()
            else:
                texto = "Tipo de arquivo não suportado."
            
            return {"text": texto, "filename": file.filename}
        finally:
            # Remover arquivo temporário
            os.unlink(temp_path)
            
    except Exception as e:
        print(f"Erro ao processar arquivo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

# Nova rota para upload de arquivos (compatível com o frontend atualizado)
@app.post("/upload-file")
async def upload_file(file: UploadFile = File(...)):
    try:
        # Verificar se as bibliotecas necessárias estão disponíveis
        if not LIBS_DISPONÍVEIS:
            raise HTTPException(
                status_code=500, 
                detail="Bibliotecas para processamento de arquivos não estão instaladas no servidor."
            )
        
        # Verificar tipo de arquivo
        tipo_arquivo = file.content_type
        tipos_permitidos = [
            "application/pdf",                                              # PDF
            "application/msword",                                           # DOC
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document", # DOCX
            "text/plain",                                                   # TXT
            "application/vnd.ms-excel",                                     # XLS
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"       # XLSX
        ]
        
        if tipo_arquivo not in tipos_permitidos:
            raise HTTPException(
                status_code=400, 
                detail=f"Tipo de arquivo não suportado: {tipo_arquivo}. Formatos aceitos: PDF, DOC, DOCX, TXT, XLS, XLSX."
            )
        
        # Verificar tamanho do arquivo (limite de 10MB)
        contents = await file.read()
        if len(contents) > 10 * 1024 * 1024:  # 10MB
            raise HTTPException(
                status_code=400, 
                detail="Arquivo muito grande. O limite é de 10MB."
            )
        
        # Extrair extensão do arquivo
        extensao = pathlib.Path(file.filename).suffix.lower()
        extensoes_permitidas = [".pdf", ".doc", ".docx", ".txt", ".xls", ".xlsx"]
        
        if extensao not in extensoes_permitidas:
            raise HTTPException(
                status_code=400, 
                detail=f"Extensão de arquivo não permitida: {extensao}. Extensões aceitas: .pdf, .doc, .docx, .txt, .xls, .xlsx"
            )
        
        # Criar diretório de upload temporário
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        
        # Gerar nome de arquivo único para evitar colisões
        nome_arquivo_seguro = f"{uuid.uuid4()}{extensao}"
        caminho_arquivo = os.path.join(UPLOAD_DIR, nome_arquivo_seguro)
        
        # Gravar arquivo temporariamente
        with open(caminho_arquivo, "wb") as f:
            f.write(contents)
        
        try:
            # Resetar o arquivo lido para a posição inicial
            file.file = io.BytesIO(contents)
            
            # Processar o arquivo baseado no tipo
            texto = ""
            if tipo_arquivo == "application/pdf":
                with open(caminho_arquivo, "rb") as f:
                    texto = extrair_texto_pdf(f)
            elif tipo_arquivo in ["application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
                texto = extrair_texto_docx(caminho_arquivo)
            elif tipo_arquivo in ["application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                texto = extrair_texto_xlsx(caminho_arquivo)
            elif tipo_arquivo == "text/plain":
                with open(caminho_arquivo, "r", encoding="utf-8") as f:
                    texto = f.read()
            
            # Limitar o tamanho do texto retornado
            max_chars = 15000  # Limitar para ~15k caracteres
            if len(texto) > max_chars:
                texto = texto[:max_chars] + f"\n\n[Texto truncado devido ao tamanho. O documento original contém {len(texto)} caracteres.]"
            
            # Registrar log de processamento bem-sucedido
            print(f"Arquivo processado com sucesso: {file.filename} ({tipo_arquivo}, {len(contents)} bytes)")
            
            return {
                "text": texto,
                "filename": file.filename,
                "size": len(contents),
                "mimetype": tipo_arquivo
            }
            
        finally:
            # Remover arquivo temporário após o processamento
            if os.path.exists(caminho_arquivo):
                os.unlink(caminho_arquivo)
                print(f"Arquivo temporário removido: {caminho_arquivo}")
    
    except HTTPException as e:
        # Repassar exceções HTTP já formatadas
        raise e
    except Exception as e:
        print(f"Erro ao processar arquivo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

# Função para gerar respostas via Ollama com streaming
async def gerar_resposta_ollama_stream(prompt, is_thinkkn_mode=False):
    try:
        # Sempre iniciar com <think> para modo de pensamento
        if is_thinkkn_mode:
            yield json.dumps({"token": "<think>"}) + "\n"
            
        # Configurar payload com parâmetros otimizados
        payload = {
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": True,  # Habilitando streaming
            "options": {
                "num_ctx": 4096,           # Contexto padrão
                "num_thread": 8,           # Usar 8 threads para paralelização
                "temperature": 0.7,        # Temperatura padrão
                "top_k": 40,               # Valor padrão
                "top_p": 0.9,              # Valor padrão
                "num_predict": 2048,       # Aumentando para respostas mais completas
                "repeat_penalty": 1.1,     # Penalidade leve para repetição
                "stop": ["<|end|>", "<|user|>"]  # Tokens de fim específicos para DeepSeek
            }
        }
        
        # IMPORTANTE: Esta função deve retornar um AsyncGenerator
        async with aiohttp.ClientSession() as session:
            try:
                print(f"Iniciando streaming com o modelo {OLLAMA_MODEL}")
                
                # Usar timeout maior para modelos mais complexos
                timeout = aiohttp.ClientTimeout(total=120)  # 120 segundos
                
                async with session.post(
                    OLLAMA_API_URL, 
                    json=payload, 
                    timeout=timeout
                ) as response:
                    if response.status != 200:
                        error_text = await response.text()
                        print(f"Erro na API do Ollama: {response.status} - {error_text}")
                        yield json.dumps({"error": f"Erro na API do Ollama: {response.status}"}) + "\n"
                        return
                    
                    # Processar o streaming de forma mais robusta
                    buffer = ""
                    think_mode = is_thinkkn_mode
                    response_started = False
                    
                    async for chunk in response.content:
                        if chunk:
                            buffer += chunk.decode('utf-8')
                            lines = buffer.split('\n')
                            
                            # Processar todas as linhas completas
                            for i in range(len(lines) - 1):
                                line = lines[i].strip()
                                if line:
                                    try:
                                        # Cada linha é um objeto JSON
                                        json_line = json.loads(line)
                                        if 'response' in json_line:
                                            token = json_line['response']
                                            
                                            # Detectar fim do pensamento e início da resposta
                                            if think_mode and "</think>" in token:
                                                think_mode = False
                                                response_started = True
                                            
                                            yield json.dumps({"token": token}) + "\n"
                                            
                                    except json.JSONDecodeError:
                                        print(f"Erro ao decodificar JSON: {line}")
                                        yield json.dumps({"error": "Erro ao decodificar resposta"}) + "\n"
                            
                            # Manter o restante no buffer
                            buffer = lines[-1]
                    
                    # Processar qualquer conteúdo restante no buffer
                    if buffer.strip():
                        try:
                            json_line = json.loads(buffer)
                            if 'response' in json_line:
                                yield json.dumps({"token": json_line['response']}) + "\n"
                        except json.JSONDecodeError:
                            pass
                    
                    # Garantir que o pensamento seja fechado corretamente
                    if think_mode:
                        yield json.dumps({"token": "</think>"}) + "\n"
                
                print("Streaming concluído com sucesso")
                
            except asyncio.TimeoutError:
                print("Timeout excedido no streaming")
                if is_thinkkn_mode:
                    yield json.dumps({"token": "</think>"}) + "\n"
                yield json.dumps({"error": "Tempo limite excedido. O modelo está demorando mais do que o esperado para responder. Por favor, tente uma pergunta mais simples."}) + "\n"
            except Exception as e:
                print(f"Erro durante o streaming: {str(e)}")
                if is_thinkkn_mode:
                    yield json.dumps({"token": "</think>"}) + "\n"
                yield json.dumps({"error": f"Erro: {str(e)}"}) + "\n"
    
    except Exception as e:
        print(f"Erro ao configurar streaming: {str(e)}")
        if is_thinkkn_mode:
            yield json.dumps({"token": "</think>"}) + "\n"
        yield json.dumps({"error": f"Erro: {str(e)}"}) + "\n"

# Rota que envia a pergunta para o modelo com streaming
@app.post("/perguntar_stream")
async def perguntar_stream(p: Pergunta):
    try:
        # Verificar se o modo Thinkkn está ativo
        is_thinkkn_mode = p.thinkknMode
        print(f"Modo Thinkkn no backend: {is_thinkkn_mode}")
        
        # Instrução específica para o modelo DeepSeek com base no modo
        if is_thinkkn_mode:
            print("Aplicando prompt específico para modo Thinkkn")
            system_prompt = """Você é um assistente policial e jurídico brasileiro especializado no sistema legal do Brasil.

FORMATO OBRIGATÓRIO: 
1. Inicie seu raciocínio SEMPRE com a tag <think>
2. Desenvolva seu raciocínio detalhadamente
3. Encerre seu raciocínio com a tag </think>
4. Após isso, apresente sua resposta final de forma clara e direta

Use terminologia técnica adequada ao contexto jurídico brasileiro. Responda sempre em português correto."""
        else:
            print("Aplicando prompt padrão (sem modo Thinkkn)")
            system_prompt = "Você é um assistente policial e jurídico brasileiro especializado no sistema legal do Brasil. Responda de forma clara, direta e em português correto."
        
        # Formato validado para DeepSeek-R1
        prompt_completo = f"<|system|>\n{system_prompt}\n<|user|>\n{p.pergunta}\n<|assistant|>\n"
        
        # Retorna um streaming response
        print(f"Iniciando streaming da resposta")
        generator = gerar_resposta_ollama_stream(prompt_completo, is_thinkkn_mode)
        return StreamingResponse(
            generator,
            media_type="application/x-ndjson"
        )
    except Exception as e:
        print(f"Erro ao processar pergunta com streaming: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar sua pergunta: {str(e)}")

# Rota que envia a pergunta via método síncrono
@app.post("/perguntar")
async def perguntar(p: Pergunta):
    try:
        # Verificar se o modo Thinkkn está ativo
        is_thinkkn_mode = p.thinkknMode
        print(f"Modo Thinkkn: {is_thinkkn_mode}")
        
        # Instrução específica para o modelo DeepSeek com base no modo
        if is_thinkkn_mode:
            system_prompt = """Você é um assistente policial e jurídico brasileiro especializado no sistema legal do Brasil.

FORMATO OBRIGATÓRIO: 
1. Inicie seu raciocínio SEMPRE com a tag <think>
2. Desenvolva seu raciocínio detalhadamente
3. Encerre seu raciocínio com a tag </think>
4. Após isso, apresente sua resposta final de forma clara e direta

Use terminologia técnica adequada ao contexto jurídico brasileiro. Responda sempre em português correto."""
        else:
            system_prompt = "Você é um assistente policial e jurídico brasileiro especializado no sistema legal do Brasil. Responda de forma clara, direta e em português correto."
        
        # Formato validado para DeepSeek-R1
        prompt_completo = f"<|system|>\n{system_prompt}\n<|user|>\n{p.pergunta}\n<|assistant|>\n"
        
        # Gera resposta via Ollama
        resposta = gerar_resposta_ollama(prompt_completo)
        
        return {"resposta": resposta}
    except Exception as e:
        print(f"Erro ao processar pergunta: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar sua pergunta: {str(e)}")

# Rota que envia a pergunta via /api/chat (para compatibilidade com o frontend)
@app.post("/api/chat")
async def chat_api(p: Pergunta):
    try:
        # Verificar explicitamente se o modo Thinkkn está ativo
        is_thinkkn_mode = p.thinkknMode
        print(f"[API /api/chat] Recebido modo Thinkkn: {is_thinkkn_mode}")
        
        # Chamar a função principal de processamento
        return await perguntar(p)
    except Exception as e:
        print(f"Erro no endpoint /api/chat: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar sua pergunta: {str(e)}")

# Rota para depuração
@app.post("/api/debug")
async def debug_api(p: Pergunta):
    try:
        # Verificar e logar o modo Thinkkn
        is_thinkkn_mode = p.thinkknMode
        print(f"[DEBUG API] Modo Thinkkn recebido: {is_thinkkn_mode}")
        
        # Retornar informações de diagnóstico
        return {
            "success": True,
            "received_params": {
                "pergunta": p.pergunta[:50] + "..." if len(p.pergunta) > 50 else p.pergunta,
                "thinkknMode": is_thinkkn_mode
            },
            "model_info": {
                "name": OLLAMA_MODEL,
                "api_url": OLLAMA_API_URL
            }
        }
    except Exception as e:
        print(f"Erro no endpoint de debug: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro no debug: {str(e)}")

# Nova rota que o frontend está tentando chamar
@app.post("/api/chat/ai")
async def api_chat_ai(request: Request):
    try:
        # Receber o corpo da requisição
        body = await request.json()
        
        # Criar um objeto Pergunta a partir do corpo
        p = Pergunta(
            pergunta=body.get("prompt", ""),
            thinkknMode=body.get("thinkknMode", False)
        )
        
        print(f"Requisição recebida em /api/chat/ai - redirecionando para perguntar_stream")
        
        # Simplesmente redirecionar para o endpoint de streaming
        return await perguntar_stream(p)
    except Exception as e:
        print(f"Erro no endpoint /api/chat/ai: {str(e)}")
        # Retornar uma resposta padrão em caso de erro
        return StreamingResponse(
            (json.dumps({"token": f"Erro ao processar sua solicitação: {str(e)}"}) + "\n" for _ in range(1)),
            media_type="application/x-ndjson"
        )

# Rota para buscar histórico de chats (implementação temporária)
@app.get("/api/chat/get")
async def get_chats():
    return []

# Iniciar o servidor se este arquivo for executado diretamente
if __name__ == "__main__":
    import uvicorn
    print("Iniciando servidor API GPTPOL na porta 8000...")
    uvicorn.run("main_api:app", host="0.0.0.0", port=8000, reload=True)